#!/usr/bin/env python3
"""
update_dashboard.py
===================
Extrae datos de un Excel FCST (hoja "Datos HTML" + "Ventas 2025")
y actualiza todas las variables JS del dashboard index.html.

Uso:
    python3 update_dashboard.py <excel.xlsx> <template.html> [output.html]

Si no se especifica output.html, sobreescribe el template.
"""

import sys
import json
import re
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: Instala openpyxl primero:  pip install openpyxl")


# ── Constantes ────────────────────────────────────────────────────────────────
RETAIL_MAP_DH = {
    'ATT': 'AT&T', 'MacStore': 'MacStore', 'iShop': 'iShop',
    'Liverpool': 'Liverpool', 'Coppel': 'Coppel', 'Mercado Libre': 'MercadoLibre',
}
RETAIL_MAP_25 = {
    'ATT': 'AT&T', 'MacStore': 'MacStore', 'iShop': 'iShop',
    'Liverpool': 'Liverpool', 'Coppel': 'Coppel',
    'Mercado Libre': 'MercadoLibre', 'MercadoLibre': 'MercadoLibre',
}
ALL_RETAILS = ['AT&T', 'MacStore', 'iShop', 'Liverpool', 'Coppel', 'MercadoLibre']


# ── Helpers ───────────────────────────────────────────────────────────────────
def fl(v):
    """Convierte a float seguro."""
    try:
        return float(v) if v not in (None, '') else 0.0
    except (TypeError, ValueError):
        return 0.0


def quarterly(monthly_dict):
    """Suma mensual → trimestral [Q1, Q2, Q3, Q4]."""
    result = {}
    for r in ALL_RETAILS:
        m = monthly_dict.get(r, [0] * 12)
        result[r] = [round(sum(m[0:3])), round(sum(m[3:6])),
                     round(sum(m[6:9])), round(sum(m[9:12]))]
    return result


def js_retail_obj(d):
    """Genera objeto JS con todos los retails."""
    parts = []
    for r in ALL_RETAILS:
        v = [round(x) for x in d.get(r, [0] * 12)]
        parts.append(f'"{r}":{json.dumps(v)}')
    return '{' + ','.join(parts) + '}'


def js_quarter_obj(q):
    parts = []
    for r in ALL_RETAILS:
        v = q.get(r, [0, 0, 0, 0])
        parts.append(f'"{r}":{json.dumps(v)}')
    return '{' + ','.join(parts) + '}'


# ── Extracción: Datos HTML (2026 FCST) ───────────────────────────────────────
def extract_datos_html(ws):
    """
    Columnas (1-indexed):
      9  = Retail
      6  = Categoría
      1  = Marca
      15-26 = FCST piezas Jan-Dec
      27-38 = Valor MXN Jan-Dec
    """
    fcst_val  = defaultdict(lambda: [0.0] * 12)
    fcst_pcs  = defaultdict(lambda: [0.0] * 12)
    marca_raw = defaultdict(lambda: defaultdict(lambda: defaultdict(
        lambda: {'val': [0.0] * 12, 'pcs': [0.0] * 12}
    )))
    cat_data  = {r: {} for r in ALL_RETAILS}
    all_brands = set()

    for row in ws.iter_rows(min_row=3, values_only=True):
        retail_raw = str(row[8]).strip() if row[8] else ''
        retail = RETAIL_MAP_DH.get(retail_raw)
        if not retail:
            continue

        brand = str(row[0]).strip() if row[0] else ''
        if not brand:
            continue
        if brand.upper() == 'TECH21':
            brand = 'Tech21'
        cat = str(row[5]).strip() if row[5] else 'Other'
        all_brands.add(brand)

        for i in range(12):
            fp = fl(row[14 + i])   # FCST piezas
            fv = fl(row[26 + i])   # Valor MXN

            fcst_pcs[retail][i] += fp
            fcst_val[retail][i] += fv

            marca_raw[brand][retail][cat]['val'][i] += fv
            marca_raw[brand][retail][cat]['pcs'][i] += fp

            if cat not in cat_data[retail]:
                cat_data[retail][cat] = {'val': [0.0] * 12, 'pcs': [0.0] * 12}
            cat_data[retail][cat]['val'][i] += fv
            cat_data[retail][cat]['pcs'][i] += fp

    return fcst_val, fcst_pcs, marca_raw, cat_data, sorted(all_brands)


# ── Extracción: Ventas 2025 ───────────────────────────────────────────────────
def extract_ventas_2025(ws):
    """
    Columnas (1-indexed):
      1  = Customer (retail)
      9-20  = piezas Jan-Dec
      23-34 = valor MXN Jan-Dec
    """
    actual_val = defaultdict(lambda: [0.0] * 12)
    actual_pcs = defaultdict(lambda: [0.0] * 12)

    for row in ws.iter_rows(min_row=3, values_only=True):
        retail_raw = str(row[0]).strip() if row[0] else ''
        retail = RETAIL_MAP_25.get(retail_raw)
        if not retail:
            continue
        for i in range(12):
            pv = fl(row[8 + i])    # piezas
            vv = fl(row[22 + i])   # valor MXN
            actual_pcs[retail][i] += pv
            actual_val[retail][i] += vv

    return actual_val, actual_pcs


# ── Extracción: Consenso → sinMovData (SKUs sin movimiento) ──────────────────
def extract_sin_movimiento_consenso(ws):
    """
    Hoja Consenso — columnas (0-indexed):
      0  = Marca
      1  = SKU Retail
      4  = Descripción
      5  = Categoría
      6  = Sub Categoría
      7  = Clase
      8  = Retail  ('ATT', 'MacStore', 'iShop', 'Liverpool', 'Coppel')
      11 = Cost unitario (landed)
      12 = Status ACA
      14 = OH Retail (unidades)
      15 = OH ACA
      20-31 = FCST Jan-Dec (piezas)
      44 = TTL FCST
      45-56 = Ventas reales Jan-Dec MXN
      58 = TTL OH $ Retail  (total ya multiplicado)
      59 = TTL OH $ Landed
      60 = TTL OH $ USD

    Condición sin-movimiento: OH Retail > 0  AND  suma(FCST Jan-Dec) == 0
    FCST/mes: promedio de ventas reales Jun (col 50) y Jul (col 51)
    """
    RETAIL_CONSENSO = {'ATT', 'MacStore', 'iShop', 'Liverpool', 'Coppel'}
    records = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 30:
            continue

        retail = str(row[8]).strip() if row[8] else ''
        if retail not in RETAIL_CONSENSO:
            continue

        oh_retail = fl(row[14])
        if oh_retail <= 0:
            continue

        # Sin movimiento = sin FCST para meses FUTUROS (Ago-Dic)
        # V27 Consenso: col25=FCST Aug, col26=Sep, col27=Oct, col28=Nov, col29=Dec
        fcst_future = sum(fl(row[25 + i]) for i in range(5))  # Aug Sep Oct Nov Dec
        if fcst_future > 0:
            continue  # tiene plan de venta futuro → no es sin movimiento

        # V27: col19=Jan..col24=Jun, col25=Aug..col29=Dec (sin Jul)
        fcst_months = [fl(row[19 + i]) for i in range(6)] + [0] + [fl(row[25 + i]) for i in range(5)]

        marca   = str(row[0]).strip()  if row[0]  else ''
        sku     = str(row[1]).strip()  if row[1]  else ''
        desc    = str(row[4]).strip()  if row[4]  else ''
        cat     = str(row[5]).strip()  if row[5]  else ''
        subcat  = str(row[6]).strip()  if row[6]  else ''
        clase   = str(row[7]).strip()  if row[7]  else ''
        status  = str(row[12]).strip() if row[12] else ''
        oh_aca  = fl(row[15])

        # FCST/mes en PIEZAS: col24=FCST Jun, col25=FCST Aug (Jul no existe en V27)
        venta_jun = fl(row[24])  # FCST Jun pcs (ventas reales Jun)
        venta_ago = fl(row[25])  # FCST Aug pcs (más reciente disponible)
        avg_fcst  = round((venta_jun + venta_ago) / 2, 2) if (venta_jun + venta_ago) > 0 else venta_jun

        # MOI: OH Retail / avg ventas mensual (si hay venta previa)
        moi = round(oh_retail / avg_fcst, 1) if avg_fcst > 0 else None

        # Costos totales: costo unitario (col11) × OH Retail (col14)
        cost_unit    = fl(row[11])
        costo_retail = round(cost_unit * oh_retail, 2)
        costo_landed = costo_retail  # misma base MXN (no hay separación landed en V27)
        costo_usd    = 0.0

        # val_riesgo = valor total OH al costo
        val_riesgo = costo_retail

        # FCST Jul-Dec para columnas de la tabla
        # V27: no hay Jul → 0; Aug=col25, Sep=col26, Oct=col27, Nov=col28, Dec=col29
        fcst_jul = 0.0; fcst_aug = fl(row[25]); fcst_sep = fl(row[26])
        fcst_oct = fl(row[27]); fcst_nov = fl(row[28]); fcst_dec = fl(row[29])

        records.append({
            'marca':        marca,
            'sku':          sku,
            'desc':         desc,
            'cat':          cat,
            'subcat':       subcat,
            'clase':        clase,
            'cliente':      retail,
            'cost':         round(fl(row[11]), 2),
            'status_aca':   status,
            'oh_retail':    int(oh_retail),
            'oh_aca':       int(oh_aca),
            'fcst_jul':     int(fcst_jul),
            'fcst_aug':     int(fcst_aug),
            'fcst_sep':     int(fcst_sep),
            'fcst_oct':     int(fcst_oct),
            'fcst_nov':     int(fcst_nov),
            'fcst_dec':     int(fcst_dec),
            'ttl_fcst':     int(sum(fcst_months)),
            'moi':          moi,
            'sin_mov':      avg_fcst,
            'val_riesgo':   round(val_riesgo, 2),
            'costo_retail': costo_retail,
            'costo_landed': costo_landed,
            'costo_usd':    costo_usd,
        })

    return records




# ── Extracción: Budget 2026 (hoja "Budget 120M") ─────────────────────────────
def extract_budget_2026(ws):
    """
    Hoja 'Budget 120M' (SKU-level, min_row=4):
      col0  = Marca
      col9  = Retail  ('ATT','MacStore','iShop','Liverpool','Coppel')
      col23-34 = Budget piezas Jan-Dec
      col36-47 = Budget $ MXN Jan-Dec

    Devuelve:
      budget_val  {retail: [12 meses MXN]}
      budget_pcs  {retail: [12 meses pzas]}
      budget_marca {marca: {retail: {val:[12], pcs:[12]}}}
    """
    RMAP = {'ATT':'AT&T','iShop':'iShop','Liverpool':'Liverpool',
            'MacStore':'MacStore','Coppel':'Coppel','MercadoLibre':'MercadoLibre'}
    BRAND_NORM = {'TECH21':'Tech21','HYPERGEAR':'Hypergear'}

    def _new():
        return {'val':[0.0]*12,'pcs':[0.0]*12}

    bv = defaultdict(lambda: [0.0]*12)
    bp = defaultdict(lambda: [0.0]*12)
    # marca → retail → {val, pcs}
    bm   = defaultdict(lambda: defaultdict(_new))
    # cat → retail → {val, pcs}
    bc   = defaultdict(lambda: defaultdict(_new))
    # marca → cat → retail → {val, pcs}
    bmc  = defaultdict(lambda: defaultdict(lambda: defaultdict(_new)))

    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or len(row) < 48 or row[9] is None:
            continue
        retail_key = str(row[9]).strip()
        retail = RMAP.get(retail_key)
        if not retail:
            continue
        marca_raw = str(row[0]).strip() if row[0] else ''
        marca = BRAND_NORM.get(marca_raw.upper(), marca_raw)
        cat   = str(row[5]).strip() if row[5] else 'Others'
        for i in range(12):
            p = fl(row[23 + i]); v = fl(row[36 + i])
            bp[retail][i] += p;           bv[retail][i] += v
            bm[marca][retail]['pcs'][i]         += p
            bm[marca][retail]['val'][i]         += v
            bc[cat][retail]['pcs'][i]            += p
            bc[cat][retail]['val'][i]            += v
            bmc[marca][cat][retail]['pcs'][i]    += p
            bmc[marca][cat][retail]['val'][i]    += v

    def _round(d):
        return {r: {'val':[round(x) for x in e['val']],'pcs':[int(x) for x in e['pcs']]}
                for r, e in d.items()}

    budget_val      = {r: [round(v) for v in vals] for r, vals in bv.items()}
    budget_pcs      = {r: [int(v)   for v in vals] for r, vals in bp.items()}
    budget_marca    = {m: _round(rd)  for m, rd  in bm.items()}
    budget_cat      = {c: _round(rd)  for c, rd  in bc.items()}
    budget_marca_cat = {m: {c: _round(rd) for c, rd in cats.items()}
                        for m, cats in bmc.items()}
    return budget_val, budget_pcs, budget_marca, budget_cat, budget_marca_cat


def build_budget_js(budget_val, budget_pcs, budget_marca, budget_cat, budget_marca_cat):
    import json as _json
    line1 = 'const budgetVal = '         + _json.dumps(budget_val,       ensure_ascii=False, separators=(',',':')) + ';'
    line2 = 'const budgetPcs = '         + _json.dumps(budget_pcs,       ensure_ascii=False, separators=(',',':')) + ';'
    line3 = 'const budgetMarcaRaw = '    + _json.dumps(budget_marca,     ensure_ascii=False, separators=(',',':')) + ';'
    line4 = 'const budgetCatRaw = '      + _json.dumps(budget_cat,       ensure_ascii=False, separators=(',',':')) + ';'
    line5 = 'const budgetMarcaCatRaw = ' + _json.dumps(budget_marca_cat, ensure_ascii=False, separators=(',',':')) + ';'
    return '\n'.join([line1, line2, line3, line4, line5])


# ── Extracción: Todos los SKUs con OH Retail (para tabla MOI completa) ────────
def extract_all_skus_with_oh(ws):
    """
    Todos los SKUs con OH Retail > 0 — sin filtro de FCST.
    Columnas Consenso (0-indexed): ver extract_sin_movimiento_consenso.
    """
    RETAIL_CONSENSO = {'ATT', 'MacStore', 'iShop', 'Liverpool', 'Coppel'}
    records = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 30:
            continue
        retail = str(row[8]).strip() if row[8] else ''
        if retail not in RETAIL_CONSENSO:
            continue
        oh_retail = fl(row[14])
        if oh_retail <= 0:
            continue

        marca  = str(row[0]).strip() if row[0] else ''
        sku    = str(row[1]).strip() if row[1] else ''
        desc   = str(row[4]).strip() if row[4] else ''
        cat    = str(row[5]).strip() if row[5] else ''
        oh_aca = fl(row[15])

        # V27: col24=FCST Jun, col25=FCST Aug (sin Jul)
        venta_jun = fl(row[24])
        venta_ago = fl(row[25])
        avg_pcs   = round((venta_jun + venta_ago) / 2, 1) if (venta_jun + venta_ago) > 0 else round(venta_jun, 1)
        moi       = round(oh_retail / avg_pcs, 1) if avg_pcs > 0 else None

        # Costo = costo unitario (col11) × OH Retail
        costo_retail = round(fl(row[11]) * oh_retail, 2)

        records.append({
            'cliente':      retail,
            'cat':          cat,
            'marca':        marca,
            'sku':          sku,
            'desc':         desc,
            'oh_retail':    int(oh_retail),
            'oh_aca':       int(oh_aca),
            'avg_pcs':      avg_pcs,
            'moi':          moi,
            'costo_retail': costo_retail,
        })
    return records


def build_all_sku_js(records):
    return 'const allSkuData = ' + json.dumps(records, ensure_ascii=False, separators=(',', ':')) + ';'

def build_sin_mov_js(records):
    """Genera la línea const sinMovData = [...]; lista de registros."""
    return 'const sinMovData = ' + json.dumps(records, ensure_ascii=False, separators=(',', ':')) + ';'


# ── Extracción: moiDataByClient (OH Retail por retail+categoría) ──────────────
RETAILS_CONSENSO = ['ATT', 'MacStore', 'iShop', 'Liverpool', 'Coppel']

def extract_moi_data_by_client(ws_consenso, ws_datos_html):
    """
    Consolida OH Retail y FCST mensual por retail+categoría para moiDataByClient.

    Consenso:
      Col 8  = Retail, Col 5 = Categoría
      Col 14 = OH Retail (pzas), Col 15 = OH ACA (pzas)
      Col 58 = TTL OH $ Retail, Col 59 = TTL OH $ Landed

    Datos HTML:
      Col 8  = Retail, Col 5 = Categoría
      Cols 14-25 = FCST pzas Jan-Dec, Cols 26-37 = FCST valor Jan-Dec

    monthly_pcs/val = promedio Jul-Dic FCST (índices 6-11).
    """
    RETAIL_MAP_DH = {
        'ATT': 'ATT', 'MacStore': 'MacStore', 'iShop': 'iShop',
        'Liverpool': 'Liverpool', 'Coppel': 'Coppel',
    }

    # Acumular OH desde Consenso — por cat y por brand
    oh_cat   = defaultdict(lambda: {'oh_retail': 0.0, 'oh_aca': 0.0,
                                    'oh_retail_val': 0.0, 'oh_aca_val': 0.0})
    oh_brand = defaultdict(lambda: {'oh_retail': 0.0, 'oh_aca': 0.0,
                                    'oh_retail_val': 0.0, 'oh_aca_val': 0.0})
    for row in ws_consenso.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 30:
            continue
        retail = str(row[8]).strip() if row[8] else ''
        if retail not in RETAILS_CONSENSO:
            continue
        cat   = str(row[5]).strip() if row[5] else 'Other'
        brand = str(row[0]).strip() if row[0] else 'Other'
        oh_r = fl(row[14]); oh_a = fl(row[15])
        # V27: costo unitario (col11) × unidades = valor total MXN
        cost_unit = fl(row[11])
        rv58 = cost_unit * oh_r   # OH Retail valor MXN
        rv59 = cost_unit * oh_a   # OH ACA valor MXN
        for r in [retail, 'all']:
            oh_cat[(r, cat)]['oh_retail']     += oh_r
            oh_cat[(r, cat)]['oh_aca']        += oh_a
            oh_cat[(r, cat)]['oh_retail_val'] += rv58
            oh_cat[(r, cat)]['oh_aca_val']    += rv59
            oh_brand[(r, brand)]['oh_retail']     += oh_r
            oh_brand[(r, brand)]['oh_aca']        += oh_a
            oh_brand[(r, brand)]['oh_retail_val'] += rv58
            oh_brand[(r, brand)]['oh_aca_val']    += rv59

    # Acumular FCST mensual desde Datos HTML — por cat y por brand
    fc_cat   = defaultdict(lambda: {'fcst_pcs': [0.0]*12, 'fcst_val': [0.0]*12})
    fc_brand = defaultdict(lambda: {'fcst_pcs': [0.0]*12, 'fcst_val': [0.0]*12})
    for row in ws_datos_html.iter_rows(min_row=3, values_only=True):
        retail_raw = str(row[8]).strip() if row[8] else ''
        retail = RETAIL_MAP_DH.get(retail_raw)
        if not retail:
            continue
        cat   = str(row[5]).strip() if row[5] else 'Other'
        brand = str(row[0]).strip() if row[0] else 'Other'
        if brand.upper() == 'TECH21':
            brand = 'Tech21'
        for i in range(12):
            p = fl(row[14 + i]); v = fl(row[26 + i])
            for r in [retail, 'all']:
                fc_cat[(r, cat)]['fcst_pcs'][i]   += p
                fc_cat[(r, cat)]['fcst_val'][i]   += v
                fc_brand[(r, brand)]['fcst_pcs'][i]  += p
                fc_brand[(r, brand)]['fcst_val'][i]  += v

    def _build_rows(oh_dict, fc_dict, client):
        keys = {k[1] for k in oh_dict if k[0] == client and oh_dict[k]['oh_retail'] > 0}
        rows = []
        for key in sorted(keys):
            oh  = oh_dict[(client, key)]
            fc  = fc_dict.get((client, key), {'fcst_pcs': [0.0]*12, 'fcst_val': [0.0]*12})
            pcs_h2 = fc['fcst_pcs'][6:12]; val_h2 = fc['fcst_val'][6:12]
            mpcs = sum(pcs_h2) / 6 if any(pcs_h2) else 0.0
            mval = sum(val_h2) / 6 if any(val_h2) else 0.0
            orv = oh['oh_retail_val']; oav = oh['oh_aca_val']
            orp = oh['oh_retail'];     oap = oh['oh_aca']
            rows.append({
                'key':            key,
                'oh_retail':      round(orp, 1),
                'oh_aca':         round(oap, 1),
                'oh_retail_val':  round(orv, 2),
                'oh_aca_val':     round(oav, 2),
                'monthly_pcs':    round(mpcs, 1),
                'monthly_val':    round(mval, 1),
                'moi_retail_val': round(orv / mval, 1) if mval > 0 else None,
                'moi_aca_val':    round(oav / mval, 1) if mval > 0 else None,
                'moi_retail_pcs': round(orp / mpcs, 1) if mpcs > 0 else None,
                'moi_aca_pcs':    round(oap / mpcs, 1) if mpcs > 0 else None,
            })
        rows.sort(key=lambda x: -(x['oh_retail_val'] or 0))
        return rows

    # Construir resultado por cliente con vistas 'cat' y 'brand'
    result = {}
    for client in ['all'] + RETAILS_CONSENSO:
        result[client] = {
            'cat':   _build_rows(oh_cat,   fc_cat,   client),
            'brand': _build_rows(oh_brand, fc_brand, client),
        }
    return result


def build_moi_by_client_js(data):
    return 'const moiDataByClient = ' + json.dumps(data, ensure_ascii=False, separators=(',', ':')) + ';'


# ── Genera JS de marcaRaw ─────────────────────────────────────────────────────
def build_marca_raw_js(marca_raw):
    parts_b = []
    for b in sorted(marca_raw.keys()):
        parts_r = []
        for r in ALL_RETAILS:
            if r not in marca_raw[b]:
                continue
            parts_c = []
            for c in sorted(marca_raw[b][r].keys()):
                cd = marca_raw[b][r][c]
                if sum(cd['val']) + sum(cd['pcs']) < 0.01:
                    continue
                v = [round(x) for x in cd['val']]
                p = [round(x) for x in cd['pcs']]
                parts_c.append(f'"{c}":{{"val":{json.dumps(v)},"pcs":{json.dumps(p)}}}')
            if parts_c:
                parts_r.append(f'"{r}":{{{",".join(parts_c)}}}')
        if parts_r:
            parts_b.append(f'"{b}":{{{",".join(parts_r)}}}')
    return 'const marcaRaw = {' + ','.join(parts_b) + '};'


# ── Genera JS de catData ──────────────────────────────────────────────────────
def build_cat_data_js(cat_data):
    result = {}
    for r in ALL_RETAILS:
        result[r] = {}
        for cat, cd in cat_data.get(r, {}).items():
            result[r][cat] = {
                'val': [round(v, 2) for v in cd['val']],
                'pcs': [round(p) for p in cd['pcs']],
            }
    return 'const catData = ' + json.dumps(result) + ';'


# ── Reemplaza una variable JS en el HTML ──────────────────────────────────────
def replace_js_var(html, var_name, new_line):
    lines = html.split('\n')
    new_lines = []
    replaced = 0
    for line in lines:
        if line.strip().startswith(f'const {var_name}') and '=' in line:
            indent = len(line) - len(line.lstrip())
            new_lines.append(' ' * indent + new_line)
            replaced += 1
        else:
            new_lines.append(line)
    return '\n'.join(new_lines), replaced


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    excel_path   = Path(sys.argv[1])
    html_path    = Path(sys.argv[2])
    output_path  = Path(sys.argv[3]) if len(sys.argv) >= 4 else html_path

    if not excel_path.exists():
        sys.exit(f"ERROR: No se encuentra el Excel: {excel_path}")
    if not html_path.exists():
        sys.exit(f"ERROR: No se encuentra el HTML: {html_path}")

    print(f"📂 Excel:    {excel_path.name}")
    print(f"📄 Template: {html_path.name}")
    print(f"💾 Output:   {output_path.name}")
    print()

    # Cargar Excel
    print("⏳ Cargando Excel…")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    if 'Datos HTML' not in wb.sheetnames:
        sys.exit("ERROR: No se encontró la hoja 'Datos HTML' en el Excel.")
    if 'Ventas 2025' not in wb.sheetnames:
        sys.exit("ERROR: No se encontró la hoja 'Ventas 2025' en el Excel.")

    ws_dh = wb['Datos HTML']
    ws_25 = wb['Ventas 2025']

    # ── Consenso → sinMovData (fuente principal de SKUs sin movimiento) ─────────
    sin_mov_records = []
    if 'Consenso' in wb.sheetnames:
        ws_cons = wb['Consenso']
        print("⏳ Extrayendo SKUs sin movimiento (Consenso)…")
        sin_mov_records = extract_sin_movimiento_consenso(ws_cons)
        from collections import Counter
        clients_cnt = Counter(r['cliente'] for r in sin_mov_records)
        print(f"   ✅ sinMovData: {len(sin_mov_records)} registros → {dict(clients_cnt)}")
    else:
        print("   ⚠️  Hoja 'Consenso' no encontrada — sinMovData no se actualizará")

    # Extraer datos
    print("⏳ Extrayendo datos 2026 (Datos HTML)…")
    fcst_val, fcst_pcs, marca_raw, cat_data, all_brands = extract_datos_html(ws_dh)

    print("⏳ Extrayendo datos 2025 (Ventas 2025)…")
    actual_val, actual_pcs = extract_ventas_2025(ws_25)

    # Calcular trimestrales
    qv   = quarterly(fcst_val)
    qp   = quarterly(fcst_pcs)
    q25v = quarterly(actual_val)
    q25p = quarterly(actual_pcs)

    # Validar totales
    grand_total_2026 = round(sum(sum(v) for v in fcst_val.values()))
    grand_pcs_2026   = round(sum(sum(p) for p in fcst_pcs.values()))
    grand_total_2025 = round(sum(sum(v) for v in actual_val.values()))
    print(f"   ✅ 2026 MXN total:  ${grand_total_2026:>15,.0f}")
    print(f"   ✅ 2026 pzas total: {grand_pcs_2026:>15,.0f}")
    print(f"   ✅ 2025 MXN total:  ${grand_total_2025:>15,.0f}")
    print()

    # Construir líneas JS
    vars_to_replace = {
        'avb2026Val':  f'const avb2026Val = {js_retail_obj(fcst_val)};',
        'avb2026Pcs':  f'const avb2026Pcs = {js_retail_obj(fcst_pcs)};',
        'valorFull':   f'const valorFull = {js_retail_obj(fcst_val)};',
        'piezasAnual': f'const piezasAnual = {js_retail_obj(fcst_pcs)};',
        'avb2025Val':  f'const avb2025Val = {js_retail_obj(actual_val)};',
        'avb2025Pcs':  f'const avb2025Pcs = {js_retail_obj(actual_pcs)};',
        'qValorAll':   f'const qValorAll = {js_retail_obj(qv)};',
        'qPiezas':     f'const qPiezas   = {js_retail_obj(qp)};',
        'q2025Val':    f'const q2025Val  = {js_retail_obj(q25v)};',
        'q2025Pcs':    f'const q2025Pcs  = {js_retail_obj(q25p)};',
        'ALL_BRANDS':  f'const ALL_BRANDS = {json.dumps(all_brands)};',
    }

    # Leer HTML
    print("⏳ Actualizando HTML…")
    html = html_path.read_text(encoding='utf-8')

    # Reemplazar variables simples
    for var_name, new_line in vars_to_replace.items():
        html, n = replace_js_var(html, var_name, new_line)
        status = '✅' if n > 0 else '⚠️ NOT FOUND'
        print(f"   {status} {var_name}")

    # Reemplazar marcaRaw (puede ser muy larga)
    html, n = replace_js_var(html, 'marcaRaw', build_marca_raw_js(marca_raw))
    print(f"   {'✅' if n > 0 else '⚠️ NOT FOUND'} marcaRaw ({len(all_brands)} marcas)")

    # Reemplazar catData
    html, n = replace_js_var(html, 'catData', build_cat_data_js(cat_data))
    print(f"   {'✅' if n > 0 else '⚠️ NOT FOUND'} catData")

    # Reemplazar allSkuData (todos los SKUs con OH) desde Consenso
    if 'Consenso' in wb.sheetnames:
        all_sku_records = extract_all_skus_with_oh(wb['Consenso'])
        ALL_SKU_RE = re.compile(r'const allSkuData\s*=\s*\[.*?\];', re.DOTALL)
        new_all_sku = build_all_sku_js(all_sku_records)
        html, n_ask = ALL_SKU_RE.subn(new_all_sku, html, count=1)
        print(f"   {'✅' if n_ask else '⚠️ NOT FOUND'} allSkuData ({len(all_sku_records)} SKUs totales)")

    # Reemplazar sinMovData desde Consenso (fuente autoritativa)
    if sin_mov_records:
        SINMOV_RE = re.compile(r'const sinMovData\s*=\s*\[.*?\];', re.DOTALL)
        new_sinmov = build_sin_mov_js(sin_mov_records)
        html, n_sm = SINMOV_RE.subn(new_sinmov, html, count=1)
        print(f"   {'✅' if n_sm else '⚠️ NOT FOUND'} sinMovData ({len(sin_mov_records)} SKUs desde Consenso)")

    # Reemplazar moiDataByClient desde Consenso + Datos HTML
    if 'Consenso' in wb.sheetnames:
        print("⏳ Calculando moiDataByClient (OH por retail+categoría)…")
        moi_by_client = extract_moi_data_by_client(wb['Consenso'], ws_dh)
        MOI_BC_RE = re.compile(r'const moiDataByClient\s*=\s*\{.*?\};', re.DOTALL)
        new_mbc = build_moi_by_client_js(moi_by_client)
        html, n_mbc = MOI_BC_RE.subn(new_mbc, html, count=1)
        cats_all = len(moi_by_client.get('all', {}).get('cat', []))
        print(f"   {'✅' if n_mbc else '⚠️ NOT FOUND'} moiDataByClient ({cats_all} categorías)")

    # Reemplazar budgetVal / budgetPcs / budgetMarcaRaw (hoja Budget 120M)
    budget_sheet = 'Budget 120M' if 'Budget 120M' in wb.sheetnames else 'Budget 2026'
    if budget_sheet in wb.sheetnames:
        bval, bpcs, bmarca, bcat, bmc = extract_budget_2026(wb[budget_sheet])
        BUDGET_RE = re.compile(
            r'const budgetVal\s*=\s*\{.*?\};\nconst budgetPcs\s*=\s*\{.*?\};\n'
            r'const budgetMarcaRaw\s*=\s*\{.*?\};\nconst budgetCatRaw\s*=\s*\{.*?\};\n'
            r'const budgetMarcaCatRaw\s*=\s*\{.*?\};',
            re.DOTALL
        )
        new_budget = build_budget_js(bval, bpcs, bmarca, bcat, bmc)
        html, n_b = BUDGET_RE.subn(new_budget, html, count=1)
        total_b = sum(sum(v) for v in bval.values())
        print(f"   {'✅' if n_b else '⚠️ NOT FOUND'} budget vars (${total_b/1e6:.1f}M, {len(bmarca)} marcas, {len(bcat)} cats)")
    else:
        print("   ⚠️  Hoja 'Budget 120M' no encontrada — budget no actualizado")

    # Guardar
    output_path.write_text(html, encoding='utf-8')
    print()
    print(f"✅ Dashboard actualizado → {output_path}")
    print(f"   Tamaño: {output_path.stat().st_size / 1024:.0f} KB")


if __name__ == '__main__':
    main()
